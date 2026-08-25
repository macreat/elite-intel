"""Tests for stock write-back into the source Excel catalog file."""

import openpyxl
import pytest

from app.core.config import settings
from app.models.product import Product


HEADERS = ["ARTICULO", "VALOR FACTURA", "VALOR / LOCAL", "GANANCIA", "MARGEN %", "STOCK", "Nota"]


def _make_catalog(tmp_path, rows):
    """Create a minimal Precios-sheet workbook; returns its path string."""
    path = tmp_path / "PRECIOS_PRODUCTOS_PAPELERIA.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Precios"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return str(path)


def _read_cell(path, row_num):
    wb = openpyxl.load_workbook(path)
    return wb["Precios"].cell(row=row_num, column=6).value


def _ensure_category(db_session):
    from app.models.category import Category

    cat = db_session.query(Category).first()
    if not cat:
        cat = Category(name="Test", type="EXPENSE")
        db_session.add(cat)
        db_session.commit()
        db_session.refresh(cat)
    return cat


def _make_product(db_session, name, **kwargs):
    cat = _ensure_category(db_session)
    product = Product(name=name, category_id=cat.id, **kwargs)
    db_session.add(product)
    db_session.commit()
    db_session.refresh(product)
    return product


@pytest.fixture()
def catalog_file(tmp_path, monkeypatch):
    def _factory(rows):
        path = _make_catalog(tmp_path, rows)
        monkeypatch.setattr(settings, "CATALOG_XLSX_PATH", path)
        return path

    return _factory


class TestSingleStockSync:
    def test_patch_stock_writes_value_into_xlsx(self, client, db_session, catalog_file):
        path = catalog_file([["LAPIZ NEGRO", 100, 200, None, None, 5, None]])
        product = _make_product(db_session, "LAPIZ NEGRO", stock_qty=1)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 12})
        assert res.status_code == 200
        assert _read_cell(path, 2) == 12

    def test_patch_normalizes_case_and_whitespace_for_match(self, client, db_session, catalog_file):
        path = catalog_file([["CUADERNO   A5 ", 100, 200, None, None, 3, None]])
        product = _make_product(db_session, "cuaderno a5", stock_qty=0)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 8})
        assert res.status_code == 200
        assert _read_cell(path, 2) == 8

    def test_patch_accents_preserved_by_shared_rule(self, client, db_session, catalog_file):
        path = catalog_file([["LÁPIZ NEGRO", 100, 200, None, None, 5, None]])
        product = _make_product(db_session, "lápiz negro", stock_qty=1)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 12})
        assert res.status_code == 200
        assert _read_cell(path, 2) == 12

    def test_patch_null_stock_clears_cell(self, client, db_session, catalog_file):
        path = catalog_file([["BORRADOR BLANCO", 100, 200, None, None, 4, None]])
        product = _make_product(db_session, "BORRADOR BLANCO", stock_qty=4)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": None})
        assert res.status_code == 200
        assert _read_cell(path, 2) is None

    def test_patch_other_cells_untouched(self, client, db_session, catalog_file):
        path = catalog_file([["TIJERA ORO", 500, 900, 10, "40%", 1, "nota fija"]])
        product = _make_product(db_session, "TIJERA ORO", stock_qty=1)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 21})
        assert res.status_code == 200
        wb = openpyxl.load_workbook(path)
        row = [c.value for c in wb["Precios"][2]]
        assert row == ["TIJERA ORO", 500, 900, 10, "40%", 21, "nota fija"]


class TestBulkStockSync:
    def test_bulk_updates_all_rows_with_single_save(self, client, db_session, catalog_file, monkeypatch):
        path = catalog_file(
            [
                ["BOLIGRAFO AZUL", 100, 200, None, None, 1, None],
                ["BOLIGRAFO ROJO", 100, 200, None, None, 2, None],
                ["LAPIZ MINE", 100, 200, None, None, 9, None],
            ]
        )
        p1 = _make_product(db_session, "BOLIGRAFO AZUL", stock_qty=1)
        p2 = _make_product(db_session, "BOLIGRAFO ROJO", stock_qty=2)
        p3 = _make_product(db_session, "LAPIZ MINE", stock_qty=9)

        save_calls = []
        original_save = openpyxl.workbook.workbook.Workbook.save

        def counting_save(self, filename):
            save_calls.append(filename)
            return original_save(self, filename)

        monkeypatch.setattr(openpyxl.workbook.workbook.Workbook, "save", counting_save)

        res = client.post(
            "/api/v1/products/stock/bulk",
            json={"items": [
                {"product_id": p1.id, "stock": 10},
                {"product_id": p2.id, "stock": 0},
                {"product_id": p3.id, "stock": None},
            ]},
        )
        assert res.status_code == 200
        assert len(save_calls) == 1
        assert _read_cell(path, 2) == 10
        assert _read_cell(path, 3) == 0
        assert _read_cell(path, 4) is None


class TestSyncEdgeCases:
    def test_unknown_product_name_skipped(self, client, db_session, catalog_file):
        path = catalog_file(
            [
                ["EN EL CATALOGO", 100, 200, None, None, 7, None],
                ["OTRO ITEM", 100, 200, None, None, 3, None],
            ]
        )
        product = _make_product(db_session, "SOLO EN DB", stock_qty=None)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 5})
        assert res.status_code == 200
        assert _read_cell(path, 2) == 7
        assert _read_cell(path, 3) == 3

    def test_missing_file_does_not_raise(self, client, db_session, tmp_path, monkeypatch):
        missing = str(tmp_path / "NO_EXISTE.xlsx")
        monkeypatch.setattr(settings, "CATALOG_XLSX_PATH", missing)
        product = _make_product(db_session, "REGLA 30CM", stock_qty=4)

        res = client.patch(f"/api/v1/products/{product.id}/stock", json={"stock": 15})
        assert res.status_code == 200
        assert res.json()["stock_qty"] == 15

    def test_missing_file_does_not_raise_on_bulk(self, client, db_session, tmp_path, monkeypatch):
        missing = str(tmp_path / "TAMPOCO.xlsx")
        monkeypatch.setattr(settings, "CATALOG_XLSX_PATH", missing)
        p1 = _make_product(db_session, "CARPETA", stock_qty=6)

        res = client.post(
            "/api/v1/products/stock/bulk",
            json={"items": [{"product_id": p1.id, "stock": 50}]},
        )
        assert res.status_code == 200


def _read_row(path, row_num):
    wb = openpyxl.load_workbook(path)
    return [c.value for c in wb["Precios"][row_num]]


class TestPriceSync:
    def test_patch_price_writes_invoice_and_local_columns(self, client, db_session, catalog_file):
        path = catalog_file([["LAPIZ NEGRO", 100, 200, None, None, 5, None]])
        product = _make_product(
            db_session, "LAPIZ NEGRO", invoice_price=100, local_price=200, stock_qty=5
        )

        res = client.patch(
            f"/api/v1/products/{product.id}",
            json={"invoice_price": 150.5, "local_price": 280},
        )

        assert res.status_code == 200
        assert _read_row(path, 2) == ["LAPIZ NEGRO", 150.5, 280, None, None, 5, None]

    def test_patch_single_price_leaves_other_price_column(self, client, db_session, catalog_file):
        path = catalog_file([["CUADERNO A5", 100, 200, None, None, 3, None]])
        product = _make_product(
            db_session, "CUADERNO A5", invoice_price=100, local_price=200, stock_qty=3
        )

        res = client.patch(f"/api/v1/products/{product.id}", json={"invoice_price": 120})

        assert res.status_code == 200
        assert _read_row(path, 2) == ["CUADERNO A5", 120, 200, None, None, 3, None]

    def test_bulk_prices_updates_all_rows_with_single_save(
        self, client, db_session, catalog_file, monkeypatch
    ):
        path = catalog_file(
            [
                ["BOLIGRAFO AZUL", 100, 200, None, None, 1, None],
                ["BOLIGRAFO ROJO", 100, 200, None, None, 2, None],
                ["LAPIZ MINE", 100, 200, None, None, 9, None],
            ]
        )
        p1 = _make_product(db_session, "BOLIGRAFO AZUL", invoice_price=100, local_price=200)
        p2 = _make_product(db_session, "BOLIGRAFO ROJO", invoice_price=100, local_price=200)
        p3 = _make_product(db_session, "LAPIZ MINE", invoice_price=100, local_price=200)

        save_calls = []
        original_save = openpyxl.workbook.workbook.Workbook.save

        def counting_save(self, filename):
            save_calls.append(filename)
            return original_save(self, filename)

        monkeypatch.setattr(openpyxl.workbook.workbook.Workbook, "save", counting_save)

        res = client.post(
            "/api/v1/products/prices/bulk",
            json={
                "items": [
                    {"product_id": p1.id, "invoice_price": 110, "local_price": 210},
                    {"product_id": p2.id, "invoice_price": 120},
                    {"product_id": p3.id, "local_price": 250},
                ]
            },
        )

        assert res.status_code == 200
        assert len(save_calls) == 1
        assert _read_row(path, 2) == ["BOLIGRAFO AZUL", 110, 210, None, None, 1, None]
        assert _read_row(path, 3) == ["BOLIGRAFO ROJO", 120, 200, None, None, 2, None]
        assert _read_row(path, 4) == ["LAPIZ MINE", 100, 250, None, None, 9, None]

    def test_missing_file_does_not_raise_on_price_patch(
        self, client, db_session, tmp_path, monkeypatch
    ):
        missing = str(tmp_path / "SIN_ARCHIVO.xlsx")
        monkeypatch.setattr(settings, "CATALOG_XLSX_PATH", missing)
        product = _make_product(db_session, "REGLA METALICA", invoice_price=10, local_price=20)

        res = client.patch(f"/api/v1/products/{product.id}", json={"invoice_price": 15})
        assert res.status_code == 200
        assert float(res.json()["invoice_price"]) == 15

    def test_missing_file_does_not_raise_on_bulk_prices(
        self, client, db_session, tmp_path, monkeypatch
    ):
        missing = str(tmp_path / "NADA.xlsx")
        monkeypatch.setattr(settings, "CATALOG_XLSX_PATH", missing)
        p1 = _make_product(db_session, "ESPIRAL GRANDE", invoice_price=30, local_price=60)

        res = client.post(
            "/api/v1/products/prices/bulk",
            json={"items": [{"product_id": p1.id, "invoice_price": 35}]},
        )
        assert res.status_code == 200
