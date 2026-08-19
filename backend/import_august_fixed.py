#!/usr/bin/env python3
"""Re-import August data from xlsx — use numeric values directly, fix text cell parsing."""
import hashlib, json, re
from decimal import Decimal, InvalidOperation
import openpyxl, psycopg

DB_URL = "postgresql://postgres:postgres@elite_postgres:5432/elite_db"
XLSX_PATH = "/app/data/kardex_august.xlsx"
CONTENT_HASH = hashlib.sha256(b"kardex_august_2026_v3_fixed").hexdigest()

def parse_text_amount(text):
    """Parse amount from text cells like '627.000 Arriendo' — dot here is DECIMAL (627.00), not thousands."""
    if not text: return None
    s = str(text).strip()
    s = re.sub(r'[$\s]', '', s)
    m = re.search(r'[-+]?\d[\d.,]*', s)
    if not m: return None
    raw = m.group()
    has_dot = '.' in raw
    has_comma = ',' in raw
    if has_dot and has_comma:
        # Last separator is decimal
        lp = max(raw.rfind('.'), raw.rfind(','))
        normalized = raw[:lp].replace('.','').replace(',','') + '.' + raw[lp+1:]
    elif has_comma:
        # Comma = decimal
        normalized = raw.replace(',', '.')
    elif has_dot:
        # Dot = decimal (627.000 means 627.00)
        normalized = raw
    else:
        normalized = raw
    try: v = Decimal(normalized)
    except: return None
    if v <= 0: return None
    return v.quantize(Decimal("0.01"))

def to_cop(v):
    """Convert a raw xlsx number to COP Decimal."""
    if v is None: return None
    if isinstance(v, (int, float)):
        d = Decimal(str(v))
        return d.quantize(Decimal("0.01")) if d > 0 else None
    # Text cell
    return parse_text_amount(v)

def source_fp(ch, key): return hashlib.sha256(f"{ch}:{key}".encode()).hexdigest()

def classify(h):
    h = h.lower().strip()
    if not h or h in ('pendientes','total'): return None
    if 'ahorro' in h and 'pagar' in h: return ('Ahorro para pagar','EXPENSE')
    if h in ('salida','salidas'): return ('Salidas','EXPENSE')
    if 'movil' in h: return ('Be Movil','INCOME')
    if 'tigo' in h: return ('Tigo','INCOME')
    if 'fotocopi' in h: return ('Fotocopias','INCOME')
    if 'impresion' in h: return ('Impresiones','INCOME')
    if 'scan' in h or 'escaneo' in h: return ('Escaneo','INCOME')
    if 'papeler' in h: return ('Papelería','INCOME')
    if 'accesorio' in h: return ('Accesorios','INCOME')
    if 'internet' in h: return ('Internet','INCOME')
    if 'ahorro mensual' in h or h == 'ahorro': return ('Ahorro mensual','INCOME')
    if h.replace('.','').replace(',','').isdigit(): return None
    return ('Otros','INCOME')

def main():
    conn = psycopg.connect(DB_URL)
    cur = conn.cursor()

    print("=== WIPING ===")
    cur.execute("DELETE FROM transactions")
    cur.execute("DELETE FROM import_rows")
    cur.execute("DELETE FROM import_batches")
    cur.execute("DELETE FROM products")
    cur.execute("DELETE FROM categories")
    conn.commit()

    print("=== CATEGORIES ===")
    cats = [
        ("Ahorro mensual","INCOME"),("Be Movil","INCOME"),("Tigo","INCOME"),
        ("Fotocopias","INCOME"),("Impresiones","INCOME"),("Escaneo","INCOME"),
        ("Papelería","INCOME"),("Accesorios","INCOME"),("Internet","INCOME"),
        ("Otros","INCOME"),("Recargas","INCOME"),("Servicios digitales","INCOME"),
        ("Salidas","EXPENSE"),("Ahorro para pagar","EXPENSE"),
    ]
    cat_ids = {}
    for name, tx_type in cats:
        cur.execute("INSERT INTO categories (name,type,active,created_at,updated_at) VALUES (%s,%s,true,NOW(),NOW()) RETURNING id",(name,tx_type))
        cat_ids[(name.lower(),tx_type)] = cur.fetchone()[0]
    conn.commit()
    print(f"  {len(cats)} categories")

    print("=== PRODUCTS (CATALOG sheet) ===")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['CATALOG']
    pc = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row[0] else ""
        if not name or name == 'None': continue
        def cp(v):
            if v is None: return None
            if isinstance(v, (int,float)): return Decimal(str(v)).quantize(Decimal("0.01")) if v > 0 else None
            m = re.search(r'[\d.]+', str(v).strip())
            if m:
                try: return Decimal(m.group())
                except: return None
            return None
        invoice = cp(row[1]) if len(row)>1 else None
        local = cp(row[2]) if len(row)>2 else None
        cur.execute("INSERT INTO products (name,invoice_price,local_price,currency_code,created_at,updated_at) VALUES (%s,%s,%s,'COP',NOW(),NOW())",(name, invoice, local))
        pc += 1
    wb.close()
    conn.commit()
    print(f"  {pc} products")

    print("=== TRANSACTIONS (August only, fixed amounts) ===")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['KardexTestMVP']
    inserted = skipped = errors = 0
    current_date = None
    current_header = None
    row_idx = 0
    batch = []

    for row in ws.iter_rows(values_only=True):
        row_idx += 1
        col0 = row[0]
        if col0 is not None and hasattr(col0, 'strftime'):
            current_date = col0
            continue
        if col0 is not None and isinstance(col0, str) and col0.strip().lower() in ('ahorro mensual','ahorro'):
            current_header = [str(c).strip() if c else '' for c in row]
            continue
        if not current_header or not current_date: continue
        occurred_at = current_date.strftime("%Y-%m-%dT00:00:00-05:00")

        for ci in range(1, min(len(row), len(current_header))):
            hdr = current_header[ci]
            res = classify(hdr)
            if res is None: continue
            cat_name, tx_type = res
            cat_id = cat_ids.get((cat_name.lower(), tx_type))
            if cat_id is None: continue

            val = row[ci]
            if val is None: continue

            amount = to_cop(val)
            if amount is None or amount <= 0:
                skipped += 1
                continue

            desc = str(val).strip() if isinstance(val, str) else hdr
            fp = source_fp(CONTENT_HASH, f"{row_idx}:{ci}:{amount}")
            rec = json.dumps({"occurred_at":occurred_at,"transaction_type":tx_type,"category_id":cat_id,"description":desc,"amount":str(amount),"currency_code":"COP","product_id":None},sort_keys=True)
            rec_hash = hashlib.sha256(rec.encode()).hexdigest()
            batch.append((occurred_at, tx_type, cat_id, desc, str(amount), row_idx, rec_hash, fp))

        if len(batch) >= 200:
            for b in batch:
                try:
                    cur.execute("INSERT INTO transactions (occurred_at,transaction_type,category_id,description,amount,currency_code,source_type,import_batch_id,source_row_number,record_fingerprint,source_fingerprint,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,'COP','CSV',NULL,%s,%s,%s,NOW(),NOW()) ON CONFLICT DO NOTHING",b)
                    if cur.rowcount > 0: inserted += 1
                    else: skipped += 1
                except: errors += 1
            conn.commit()
            batch = []
            print(f"  ... {inserted} (row {row_idx})")

    for b in batch:
        try:
            cur.execute("INSERT INTO transactions (occurred_at,transaction_type,category_id,description,amount,currency_code,source_type,import_batch_id,source_row_number,record_fingerprint,source_fingerprint,created_at,updated_at) VALUES (%s,%s,%s,%s,%s,'COP','CSV',NULL,%s,%s,%s,NOW(),NOW()) ON CONFLICT DO NOTHING",b)
            if cur.rowcount > 0: inserted += 1
            else: skipped += 1
        except: errors += 1
    conn.commit()
    wb.close()

    print(f"\n=== RESULTS ===")
    print(f"  Inserted: {inserted}, Skipped: {skipped}, Errors: {errors}")

    print("\n=== VERIFICATION ===")
    cur.execute("SELECT transaction_type, COUNT(*), SUM(amount) FROM transactions GROUP BY transaction_type ORDER BY transaction_type")
    for r in cur.fetchall(): print(f"  {r[0]}: {r[1]} txs, {r[2]} COP")
    cur.execute("SELECT MIN(occurred_at), MAX(occurred_at) FROM transactions")
    r = cur.fetchone()
    print(f"  Date range: {r[0]} to {r[1]}")
    cur.execute("SELECT description, amount FROM transactions ORDER BY amount DESC LIMIT 5")
    print("  Top 5 amounts:")
    for r in cur.fetchall(): print(f"    {r[1]} COP — {r[0]!r}")
    conn.close()
    print("\nDone.")

if __name__ == "__main__":
    main()
