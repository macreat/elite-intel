"""Write-back of product stock into the source Excel catalog file.

The xlsx catalog (sheet `Precios`) is the human-facing source of truth for
stock. Whenever stock changes through the API, this module mirrors those
changes back into the workbook so the file always reflects current values.

Name matching reuses the importer's normalization rule from
scripts/seed_catalog.py so both directions agree on identity.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from scripts.seed_catalog import normalize_name

logger = logging.getLogger(__name__)

SHEET_NAME = "Precios"
NAME_COLUMN = 1  # ARTICULO (1-based)
INVOICE_PRICE_COLUMN = 2  # VALOR FACTURA (1-based)
LOCAL_PRICE_COLUMN = 3  # VALOR / LOCAL (1-based)
STOCK_COLUMN = 6  # STOCK (1-based)
HEADER_ROW = 1


def _build_name_row_index(ws) -> dict[str, int]:
    row_by_name: dict[str, int] = {}
    for row_num in range(HEADER_ROW + 1, ws.max_row + 1):
        raw_name = ws.cell(row=row_num, column=NAME_COLUMN).value
        if raw_name is None or not str(raw_name).strip():
            continue
        norm = normalize_name(str(raw_name))
        # First occurrence wins, mirroring the importer's dedupe rule.
        row_by_name.setdefault(norm, row_num)
    return row_by_name


def _load_sheet(xlsx_path: Path | str):
    wb = openpyxl.load_workbook(str(xlsx_path))
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in {xlsx_path}")
    return wb, wb[SHEET_NAME]


def sync_stocks_to_catalog(xlsx_path: Path | str, products) -> dict[str, int]:
    """Write each product's stock into the catalog workbook, matched by name.

    products: iterable of objects exposing `.name` and `.stock_qty`.
    Null stock clears the target cell. Names not present in the sheet are
    skipped and counted.

    Returns {"updated": n, "skipped": m}. File-level errors propagate to the
    caller, which is expected to treat export as best-effort.
    """
    wb, ws = _load_sheet(xlsx_path)

    row_by_name = _build_name_row_index(ws)

    updated = 0
    skipped = 0
    for product in products:
        norm = normalize_name(product.name)
        row_num = row_by_name.get(norm)
        if row_num is None:
            skipped += 1
            logger.warning("catalog export: '%s' not found in %s", product.name, Path(xlsx_path).name)
            continue
        ws.cell(row=row_num, column=STOCK_COLUMN).value = (
            int(product.stock_qty) if product.stock_qty is not None else None
        )
        updated += 1

    wb.save(str(xlsx_path))
    wb.close()
    return {"updated": updated, "skipped": skipped}


def sync_prices_to_catalog(xlsx_path: Path | str, products) -> dict[str, int]:
    """Write each product's prices into columns B/C of the catalog workbook.

    products: iterable of objects exposing `.name`, `.invoice_price` and
    `.local_price`. Null prices leave the target cell untouched (the API never
    nulls prices today). Names not present in the sheet are skipped and
    counted.

    Returns {"updated": n, "skipped": m}. File-level errors propagate to the
    caller, which is expected to treat export as best-effort.
    """
    wb, ws = _load_sheet(xlsx_path)

    row_by_name = _build_name_row_index(ws)

    updated = 0
    skipped = 0
    for product in products:
        norm = normalize_name(product.name)
        row_num = row_by_name.get(norm)
        if row_num is None:
            skipped += 1
            logger.warning("catalog export: '%s' not found in %s", product.name, Path(xlsx_path).name)
            continue
        if product.invoice_price is not None:
            ws.cell(row=row_num, column=INVOICE_PRICE_COLUMN).value = float(product.invoice_price)
        if product.local_price is not None:
            ws.cell(row=row_num, column=LOCAL_PRICE_COLUMN).value = float(product.local_price)
        updated += 1

    wb.save(str(xlsx_path))
    wb.close()
    return {"updated": updated, "skipped": skipped}
