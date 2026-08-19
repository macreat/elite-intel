#!/usr/bin/env python3
"""Idempotent seed script for Kardex Hoja2 price catalog.

Reads an Excel workbook, detects the header in rows 1-3, parses product
names and prices, then upserts products by normalized name. Generates
an anomaly report for skipped or flagged rows.

Usage:
    python backend/scripts/seed_catalog.py --xlsx path/to/Kardex.xlsx
    python backend/scripts/seed_catalog.py --xlsx path/to/Kardex.xlsx --dry-run
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation

from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, sessionmaker

# ---------------------------------------------------------------------------
# Normalization (D5): strip, collapse whitespace, casefold; KEEP accents
# ---------------------------------------------------------------------------

_WS_RE = re.compile(r"\s+")


def normalize_name(raw: str) -> str:
    """Strip, collapse whitespace, casefold. Accents preserved."""
    return " ".join(raw.split()).casefold()


# ---------------------------------------------------------------------------
# Price parser (D6)
# ---------------------------------------------------------------------------

_PRICE_RE = re.compile(r"\$?\s*([\d.,]+)")
_PTE_RE = re.compile(r"PTE", re.IGNORECASE)
_PTE_SUFFIX_RE = re.compile(r"PTE\b.*$", re.IGNORECASE)


def parse_price(raw: str | None) -> tuple[Decimal | None, str | None]:
    """Parse a price cell value.

    Returns (value, flag) where flag is one of None, 'PTE', 'UNCLEAN'.
    Blank/None input returns (None, None) — silently skipped.
    """
    if raw is None:
        return None, None

    text = str(raw).strip()
    if not text:
        return None, None

    # PTE detection: flag and return NULL
    if _PTE_RE.search(text):
        return None, "PTE"

    # Extract first numeric group: strip $ and thousands commas
    m = _PRICE_RE.search(text)
    if not m:
        return None, "UNCLEAN"

    num_str = m.group(1).replace(",", "")
    try:
        value = Decimal(num_str)
    except InvalidOperation:
        return None, "UNCLEAN"

    return value, None


# ---------------------------------------------------------------------------
# Header detection (D7): scan rows 1-3 for ARTICULO / FACTURA / LOCAL
# ---------------------------------------------------------------------------

_HEADER_KEYWORDS = {"articulo", "factura", "local"}


def detect_header(rows: list[tuple]) -> int | None:
    """Return the 1-based row index containing the header, or None."""
    for idx, row in enumerate(rows[:3], start=1):
        cells = [str(c).strip().casefold() if c else "" for c in row]
        joined = " ".join(cells)
        if all(kw in joined for kw in _HEADER_KEYWORDS):
            return idx
    return None


# ---------------------------------------------------------------------------
# Column mapping for header row
# ---------------------------------------------------------------------------

def find_columns(header_row: tuple) -> tuple[int, int, int]:
    """Return 0-based column indices for ARTICULO, VALOR FACTURA, VALOR LOCAL."""
    names = [str(c).strip().casefold() if c else "" for c in header_row]
    art_col = next((i for i, n in enumerate(names) if "articulo" in n), None)
    fac_col = next((i for i, n in enumerate(names) if "factura" in n), None)
    loc_col = next((i for i, n in enumerate(names) if "local" in n), None)
    if art_col is None or fac_col is None or loc_col is None:
        raise SystemExit(
            f"ERROR: Could not map columns from header: {header_row}\n"
            f"  art_col={art_col}, fac_col={fac_col}, loc_col={loc_col}"
        )
    return art_col, fac_col, loc_col


# ---------------------------------------------------------------------------
# Main seed logic
# ---------------------------------------------------------------------------

def run_seed(xlsx_path: str, sheet_name: str, dry_run: bool) -> int:
    """Execute the seed. Returns exit code (0 always — anomalies are findings)."""
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"ERROR: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        return 1
    ws = wb[sheet_name]

    # Read rows 1-3 for header detection
    preview_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    header_row_idx = detect_header(preview_rows)
    if header_row_idx is None:
        print("ERROR: Could not detect header in rows 1-3.")
        for i, row in enumerate(preview_rows, 1):
            print(f"  Row {i}: {row}")
        return 1

    header = preview_rows[header_row_idx - 1]
    art_col, fac_col, loc_col = find_columns(header)
    print(f"Header detected at row {header_row_idx}: {header}")
    print(f"  ARTICULO=col {art_col}, VALOR FACTURA=col {fac_col}, VALOR LOCAL=col {loc_col}")

    # Collect all data rows (skip title row + header row)
    data_rows: list[tuple[int, tuple]] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        data_rows.append((row_num, row))
    wb.close()

    # Parse all rows
    parsed: list[dict] = []
    anomalies: dict[str, list[str]] = defaultdict(list)
    seen_names: dict[str, int] = {}  # normalized_name → first row_num

    for row_num, row in data_rows:
        name_raw = row[art_col] if art_col < len(row) else None
        fac_raw = row[fac_col] if fac_col < len(row) else None
        loc_raw = row[loc_col] if loc_col < len(row) else None

        if name_raw is None or not str(name_raw).strip():
            continue  # blank row — skip silently

        name_clean = str(name_raw).strip()
        norm = normalize_name(name_clean)

        # Check for PTE in name — still create product, just flag it
        has_pte_in_name = "pte" in name_clean.lower()
        if has_pte_in_name:
            anomalies["PTE"].append(f"Row {row_num}: {name_clean}")

        # Parse prices
        invoice_price, inv_flag = parse_price(fac_raw)
        local_price, loc_flag = parse_price(loc_raw)

        if inv_flag:
            anomalies[inv_flag].append(f"Row {row_num}: {name_clean} (invoice: {fac_raw})")
        if loc_flag:
            anomalies[loc_flag].append(f"Row {row_num}: {name_clean} (local: {loc_raw})")

        # BLOCK CALCO: invoice > local anomaly
        if invoice_price is not None and local_price is not None and invoice_price > local_price:
            anomalies["BLOCK CALCO"].append(
                f"Row {row_num}: {name_clean} — invoice={invoice_price} > local={local_price}"
            )

        # Collision detection
        if norm in seen_names:
            anomalies["COLLISION"].append(
                f"Row {row_num}: '{name_clean}' collides with row {seen_names[norm]}"
            )
        else:
            seen_names[norm] = row_num

        parsed.append({
            "row_num": row_num,
            "name": name_clean,
            "norm": norm,
            "invoice_price": invoice_price,
            "local_price": local_price,
        })

    # Deduplicate collisions — keep first occurrence
    seen: set[str] = set()
    deduped: list[dict] = []
    for item in parsed:
        if item["norm"] not in seen:
            seen.add(item["norm"])
            deduped.append(item)

    # Summary before DB
    print(f"\nParsed {len(deduped)} unique products from {len(data_rows)} rows")
    print(f"Anomaly sections:")
    for section in ["BLOCK CALCO", "MISSING", "PTE", "UNCLEAN", "COLLISION"]:
        items = anomalies.get(section, [])
        print(f"  {section}: {len(items)}")
        for item in items:
            print(f"    - {item}")

    if dry_run:
        print("\n[DRY RUN] No database changes made.")
        return 0

    # --- Database upsert ---
    # Import models (requires app to be importable)
    sys.path.insert(0, str(__import__("pathlib").Path(__file__).resolve().parent.parent))

    from app.db.base import Base
    from app.models.product import Product

    engine = create_engine("sqlite+pysqlite:///./elite.db", future=True)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)

    with SessionLocal() as session:
        # Load all existing products into memory
        existing = {normalize_name(p.name): p for p in session.scalars(select(Product)).all()}

        created = 0
        updated = 0
        unchanged = 0

        for item in deduped:
            norm = item["norm"]
            if norm in existing:
                product = existing[norm]
                changed = False
                if product.invoice_price != item["invoice_price"]:
                    product.invoice_price = item["invoice_price"]
                    changed = True
                if product.local_price != item["local_price"]:
                    product.local_price = item["local_price"]
                    changed = True
                if changed:
                    updated += 1
                else:
                    unchanged += 1
            else:
                product = Product(
                    name=item["name"],
                    category_id=1,  # default category
                    invoice_price=item["invoice_price"],
                    local_price=item["local_price"],
                    currency_code="COP",
                    stock_qty=None,
                )
                session.add(product)
                created += 1

        session.commit()

    total_anomalies = sum(len(v) for v in anomalies.values())
    print(f"\nSeed complete:")
    print(f"  Created: {created}")
    print(f"  Updated: {updated}")
    print(f"  Unchanged: {unchanged}")
    print(f"  Anomalies: {total_anomalies}")

    return 0


def main():
    parser = argparse.ArgumentParser(description="Seed product catalog from Kardex Hoja2")
    parser.add_argument("--xlsx", required=True, help="Path to the Kardex Excel file")
    parser.add_argument("--sheet", default="Hoja2", help="Sheet name (default: Hoja2)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report only, no DB writes")
    args = parser.parse_args()

    sys.exit(run_seed(args.xlsx, args.sheet, args.dry_run))


if __name__ == "__main__":
    main()
